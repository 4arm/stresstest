import os
import json
import time
import threading
import subprocess
from datetime import datetime
from flask import Flask, request, jsonify, send_file
import psutil
import openpyxl
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

DB_PATH = 'db.xlsx'
DATA_DIR = 'data'

if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR, exist_ok=True)

lock = threading.Lock()
current_test_thread = None
test_running = False

def init_excel_db():
    if not os.path.exists(DB_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tests"
        ws.append(["Test ID", "Filename", "Start Time", "Duration"])
        wb.save(DB_PATH)

def add_test_to_db(file_path, start_time, duration):
    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb["Tests"]
    test_id = ws.max_row
    ws.append([test_id, file_path, start_time, duration])
    wb.save(DB_PATH)

def run_stress_monitor(duration, result_file):
    global test_running
    results = []
    test_running = True
    start_time = time.time()

    # Start stress-ng subprocess
    stress_cmd = ['stress-ng', '--cpu', '4', '--timeout', f'{duration}s']
    stress_proc = subprocess.Popen(stress_cmd)

    try:
        while test_running and time.time() - start_time < duration:
            cpu = psutil.cpu_percent(interval=1)
            temps = psutil.sensors_temperatures()
            temp = 0.0
            if temps:
                for key in temps:
                    if temps[key]:
                        temp = temps[key][0].current
                        break
            results.append({
                "timestamp": datetime.now().isoformat(),
                "cpu": cpu,
                "temperature": temp
            })
    finally:
        stress_proc.terminate()
        stress_proc.wait()
        with open(result_file, 'w') as f:
            json.dump(results, f, indent=2)
        os.chmod(result_file, 0o666)
        test_running = False

@app.route('/start_test', methods=['POST'])
def start_test():
    global current_test_thread, test_running
    if test_running:
        return jsonify({"status": "error", "message": "Test already running"}), 400

    data = request.json
    duration = int(data.get('duration', 20))
    start_time_str = datetime.now().isoformat()
    filename = os.path.join(DATA_DIR, f"test_{int(time.time())}.json")

    current_test_thread = threading.Thread(target=run_stress_monitor, args=(duration, filename))
    current_test_thread.start()

    with lock:
        add_test_to_db(filename, start_time_str, duration)

    return jsonify({"status": "started", "filename": filename})

@app.route('/stop_test', methods=['POST'])
def stop_test():
    global test_running
    if not test_running:
        return jsonify({"status": "error", "message": "No test is running"}), 400
    test_running = False
    return jsonify({"status": "stopping"})

@app.route('/get_test_data')
def get_test_data():
    filename = request.args.get('filename')
    if not filename or not os.path.exists(filename):
        return jsonify({"status": "error", "message": "File not found"}), 404
    return send_file(filename, mimetype='application/json')

@app.route('/get_history')
def get_history():
    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb["Tests"]
    history = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        history.append({
            "id": row[0],
            "file": row[1],
            "start_time": row[2],
            "duration": row[3]
        })
    return jsonify(history)

if __name__ == '__main__':
    init_excel_db()
    app.run(host='0.0.0.0', port=5003)