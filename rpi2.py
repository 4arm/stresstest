from datetime import datetime
import threading
from flask import Flask, jsonify, request, Response, send_file
from flask_cors import CORS
import psutil
import subprocess
import socket
import time
import os
import json
import shutil
import pandas as pd
import glob
import openpyxl



app = Flask(__name__)
CORS(app)

hostname = socket.gethostname()
ip_address = socket.gethostbyname(hostname)
RESULT_FILE = "result.json"
rpi1 = "172.18.18.20"
rpi2 = "192.168.0.50"

prev_net_io = psutil.net_io_counters()
prev_time = time.time()
stress_running = False
stress_report = None
network_running = False
network_report = None
duration = 0
network_test_process = None
total, used, free = shutil.disk_usage("/")
cpu_test_data_store = {
    "timestamp": [],
    "cpu_usage": [],
    "temperature": [],
    "Network_speed": [],
    "disk_usage": []
}
def get_read_speed(interval=1):
    # Capture initial disk IO stats
    initial_disk_io = psutil.disk_io_counters()

    # Sleep for the interval to get updated stats
    time.sleep(interval)

    # Capture disk IO stats after the interval
    updated_disk_io = psutil.disk_io_counters()

    # Calculate read speed (bytes per second)
    read_speed = (updated_disk_io.read_bytes - initial_disk_io.read_bytes) / interval
    
    # Return read speed in bytes per second
    return read_speed

def get_write_speed(interval=1):
    # Capture initial disk IO stats
    initial_disk_io = psutil.disk_io_counters()

    # Sleep for the interval to get updated stats
    time.sleep(interval)

    # Capture disk IO stats after the interval
    updated_disk_io = psutil.disk_io_counters()

    # Calculate write speed (bytes per second)
    write_speed = (updated_disk_io.write_bytes - initial_disk_io.write_bytes) / interval
    
    # Return write speed in bytes per second
    return write_speed

def get_temperature():
    try:
        sensors = psutil.sensors_temperatures()
        for key in sensors:
            if sensors[key]:
                return sensors[key][0].current
    except Exception:
        pass
    return "Unavailable"

def get_throughput():
    if os.path.exists(RESULT_FILE):
        try:
            with open(RESULT_FILE, 'r') as f:
                result = json.load(f)
                bps = result['end']['sum_received']['bits_per_second']
                mbps = round(bps / 1_000_000, 2)
                return mbps
        except Exception as e:
            return str(e)
    else:
        return "Result file not found."

def get_network_speed():
    global prev_net_io, prev_time
    current_net_io = psutil.net_io_counters()
    current_time = time.time()
    time_diff = max(current_time - prev_time, 1)

    net_speed = round(((current_net_io.bytes_sent + current_net_io.bytes_recv) -
                       (prev_net_io.bytes_sent + prev_net_io.bytes_recv)) / (1024 * 1024 * time_diff), 2)

    prev_net_io = current_net_io
    prev_time = current_time
    return net_speed

def get_system_info():
    global prev_net_io, prev_time, stress_running, net_speed

    current_net_io = psutil.net_io_counters()
    current_time = time.time()
    time_diff = max(current_time - prev_time, 1)

    net_speed = round(((current_net_io.bytes_sent + current_net_io.bytes_recv) -
                       (prev_net_io.bytes_sent + prev_net_io.bytes_recv)) / (1024 * 1024 * time_diff), 2)

    prev_net_io = current_net_io
    prev_time = current_time
    system_info = {
        "status": "online",
        "hostname": hostname,
        "ip": ip_address,
        "cpu_usage": psutil.cpu_percent(),
        "ram_used": psutil.virtual_memory().used // (1024 * 1024),
        "temperature": get_temperature(),
        "network_speed": net_speed,
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
        "stress_running": stress_running,
        "network_running": network_running,
        "throughput": get_throughput(),
        "read_speed": get_read_speed(),
        "write_speed": get_write_speed(),
        "disk_usage": {
            "total": total // (1024 * 1024),
            "used": used // (1024 * 1024),
            "free": free // (1024 * 1024)
        }
    }
    return (system_info)

def store_cpu_test_data(duration):
    global cpu_test_data_store

    cpu_test_data_store["timestamp"].clear()
    cpu_test_data_store["cpu_usage"].clear()
    cpu_test_data_store["temperature"].clear()
    cpu_test_data_store["Network_speed"].clear()
    cpu_test_data_store["disk_usage"].clear()

    for i in range(duration-1):
        cpu_usage = psutil.cpu_percent(interval=1)
        temperature = get_temperature()
        network_speed = get_system_info()["network_speed"]
        disk_usage = psutil.disk_usage('/').percent

        cpu_test_data_store["timestamp"].append(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        cpu_test_data_store["cpu_usage"].append(cpu_usage)
        cpu_test_data_store["temperature"].append(temperature)
        cpu_test_data_store["Network_speed"].append(network_speed)
        cpu_test_data_store["disk_usage"].append(disk_usage)

@app.route('/data', methods=['GET'])
def data():
    return jsonify(get_system_info())

DB_PATH = 'db.xlsx'
DATA_DIR = 'data'

if not os.path.exists(DATA_DIR):
    os.makedirs(DATA_DIR, exist_ok=True)

lock = threading.Lock()
current_test_thread = None
test_running = False

def init_excel_db():
    if not os.path.exists(DB_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tests"
        ws.append(["Test ID", "Filename", "Start Time", "Duration"])
        wb.save(DB_PATH)

init_excel_db()

def add_test_to_db(file_path, start_time, duration):
    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb["Tests"]
    test_id = ws.max_row
    ws.append([test_id, file_path, start_time, duration])
    wb.save(DB_PATH)

def run_stress_monitor(duration, result_file):
    global test_running
    results = []
    test_running = True
    start_time = time.time()

    # Start stress-ng subprocess
    stress_cmd = ['stress-ng', '--cpu', '4', '--timeout', f'{duration}s']
    stress_proc = subprocess.Popen(stress_cmd)

    try:
        while test_running and time.time() - start_time < duration:
            cpu = psutil.cpu_percent(interval=1)
            temps = psutil.sensors_temperatures()
            temp = 0.0
            if temps:
                for key in temps:
                    if temps[key]:
                        temp = temps[key][0].current
                        break
            results.append({
                "timestamp": datetime.now().isoformat(),
                "cpu": cpu,
                "temperature": temp
            })
    finally:
        stress_proc.terminate()
        stress_proc.wait()
        with open(result_file, 'w') as f:
            json.dump(results, f, indent=2)
        os.chmod(result_file, 0o666)
        test_running = False

@app.route('/start_test', methods=['POST'])
def start_test():
    global current_test_thread, test_running
    if test_running:
        return jsonify({"status": "error", "message": "Test already running"}), 400

    data = request.json
    duration = int(data.get('duration', 20))
    start_time_str = datetime.now().isoformat()
    filename = os.path.join(DATA_DIR, f"test_{int(time.time())}.json")

    current_test_thread = threading.Thread(target=run_stress_monitor, args=(duration, filename))
    current_test_thread.start()

    with lock:
        add_test_to_db(filename, start_time_str, duration)

    return jsonify({"status": "started", "filename": filename})

@app.route('/get_test_data')
def get_test_data():
    filename = request.args.get('filename')
    if not filename or not os.path.exists(filename):
        return jsonify({"status": "error", "message": "File not found"}), 404
    return send_file(filename, mimetype='application/json')

@app.route('/get_history')
def get_cpu_history():
    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb["Tests"]
    history = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        history.append({
            "id": row[0],
            "file": row[1],
            "start_time": row[2],
            "duration": row[3]
        })
    return jsonify(history)

    
def save_stress_report_with_timestamp(report):
    # Generate timestamp and count
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    filename = f"{timestamp}_CPUtest.json"

    # Save report to unique JSON file
    with open(filename, 'w') as f:
        json.dump(report, f, indent=4)

    # Record filename and timestamp in Excel
    excel_file = "stress_records.xlsx"
    try:
        df = pd.read_excel(excel_file)
    except FileNotFoundError:
        df = pd.DataFrame(columns=["Timestamp", "Filename"])

    df.loc[len(df.index)] = [timestamp, filename]
    df.to_excel(excel_file, index=False)

    return filename


STRESS_RESULT_FILE = "stress_result.json"
def append_to_stress_result(entry):
    if os.path.exists(STRESS_RESULT_FILE):
        with open(STRESS_RESULT_FILE, 'r') as f:
            try:
                stress_result = json.load(f)
            except json.JSONDecodeError:
                stress_result = []
    else:
        stress_result = []
    stress_result.append(entry)

    with open(STRESS_RESULT_FILE, 'w') as f:
        json.dump(stress_result, f, indent=4)


@app.route('/stress_result', methods=['GET'])
def get_stress_result():
    stress_result = {"report": stress_report,
                    "data": cpu_test_data_store}
    append_to_stress_result(stress_result)
    return jsonify(stress_result)

@app.route('/stop_stress', methods=['POST'])
def stop_stress():
    global stress_running, network_running

    if not stress_running | network_running:
        return jsonify({"message": "No stress test is running"}), 400

    try:
        subprocess.run(["pkill", "-f", "stress-ng"])
        subprocess.run(["pkill", "-f", "iperf3"])
        stress_running = False
        network_running = False
        return jsonify({"message": "Stress test stopped successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/report', methods=['GET'])
def get_report():
    if stress_report:
        return jsonify(stress_report)
    return jsonify({"message": "No stress test report available"}), 404

@app.route('/get-result', methods=['GET'])
def getresult():
    return jsonify({"throughput": get_throughput()})

ALERT_FILE = "alert.json"
THRESHOLD_FILE = "thresholds.json"

def append_to_alert_log(entry):
    if os.path.exists(ALERT_FILE):
        with open(ALERT_FILE, 'r') as f:
            try:
                alerts = json.load(f)
            except json.JSONDecodeError:
                alerts = []
    else:
        alerts = []
    alerts.append(entry)

    with open(ALERT_FILE, 'w') as f:
        json.dump(alerts, f, indent=4)

def load_thresholds():
    if os.path.exists(THRESHOLD_FILE):
        with open(THRESHOLD_FILE, 'r') as f:
            try:
                thresholds = json.load(f)
            except json.JSONDecodeError:
                thresholds = {}
    else:
        thresholds = {
            "cpu_temp_threshold": 70,
            "cpu_usage_threshold": 90,
            "ram_usage_threshold": 90,
            "net_speed_threshold": 90
        }
        # Create default thresholds file if not present
        with open(THRESHOLD_FILE, 'w') as f:
            json.dump(thresholds, f, indent=4)
    return thresholds

@app.route('/alerts', methods=['GET'])
def get_alerts():
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    alerts = []
    thresholds = load_thresholds()

    # Temperature check
    cpu_temp = get_temperature()
    if cpu_temp != "Unavailable" and cpu_temp > thresholds["cpu_temp_threshold"]:
        entry = {
            "time": current_time,
            "device": rpi1,
            "type": "Temperature",
            "message": f"High CPU temperature: {cpu_temp:.1f}°C"
        }
        append_to_alert_log(entry)

    # CPU Usage check
    cpu_usage = psutil.cpu_percent(interval=1)
    if cpu_usage > thresholds["cpu_usage_threshold"]:
        entry = {
            "time": current_time,
            "device": rpi1,
            "type": "Usage",
            "message": f"High CPU usage: {cpu_usage:.1f}%"
        }
        append_to_alert_log(entry)

    # RAM Usage check
    ram_usage = psutil.virtual_memory().percent
    if ram_usage > thresholds["ram_usage_threshold"]:
        entry = {
            "time": current_time,
            "device": rpi1,
            "type": "Usage",
            "message": f"High RAM usage: {ram_usage:.1f}%"
        }
        append_to_alert_log(entry)

    # Net Speed check
    net_speed = get_network_speed()
    if net_speed > thresholds["net_speed_threshold"]:
        entry = {
            "time": current_time,
            "device": rpi1,
            "type": "Net Speed",
            "message": f"High Network Speed: {net_speed:.1f} Mbps"
        }
        append_to_alert_log(entry)

    # Read the latest alert history from the file
    if os.path.exists(ALERT_FILE):
        with open(ALERT_FILE, 'r') as f:
            try:
                alerts = json.load(f)
            except json.JSONDecodeError:
                alerts = []

    return jsonify({"alerts": alerts[-50:]}), 200

@app.route('/update-thresholds', methods=['POST'])
def update_thresholds():
    try:
        thresholds = request.json
        if not thresholds:
            return jsonify({"error": "No data received"}), 400

        # Validate expected keys
        expected_keys = {"cpu_temp_threshold", "cpu_usage_threshold", "ram_usage_threshold", "net_speed_threshold"}
        if not expected_keys.issubset(thresholds.keys()):
            return jsonify({"error": "Missing keys in request"}), 400

        with open(THRESHOLD_FILE, 'w') as f:
            json.dump(thresholds, f, indent=4)

        return jsonify({"message": "Thresholds updated successfully!"}), 200

    except Exception as e:
        print("Error in update-thresholds:", e)
        return jsonify({"error": "Internal server error"}), 500

@app.route('/get-thresholds', methods=['GET'])
def get_thresholds():
    try:
        with open(THRESHOLD_FILE, 'r') as f:
            thresholds = json.load(f)
        return jsonify(thresholds), 200
    except FileNotFoundError:
        return jsonify({
            "cpu_temp_threshold": 70,
            "cpu_usage_threshold": 90,
            "ram_usage_threshold": 90,
            "net_speed_threshold": 90
        }), 200  # Defaults if file not found
    except Exception as e:
        print("Error in get-thresholds:", e)
        return jsonify({"error": "Internal server error"}), 500

HISTORY_FILE = "history.json"
previous_stress_running = False
previous_network_running = False

def append_to_history_log(entry):
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, 'r') as f:
            try:
                history = json.load(f)
            except json.JSONDecodeError:
                history = []
    else:
        history = []
    history.append(entry)

    with open(HISTORY_FILE, 'w') as f:
        json.dump(history, f, indent=4)

@app.route('/history', methods=['GET'])
def get_history():
    global stress_running, network_running
    global previous_stress_running, previous_network_running

    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Stress test logging
    if stress_running and not previous_stress_running:
        log = {
            "time": current_time,
            "device": rpi1,
            "type": "CPU Test",
            "message": "Stress test started"
        }
        append_to_history_log(log)
    elif not stress_running and previous_stress_running:
        log = {
            "time": current_time,
            "device": rpi1,
            "type": "CPU Test",
            "message": "Stress test stopped"
        }
        append_to_history_log(log)
    previous_stress_running = stress_running

    # Network test logging
    if network_running and not previous_network_running:
        log = {
            "time": current_time,
            "device": rpi1,
            "type": "Network Test",
            "message": "Network test started"
        }
        append_to_history_log(log)
    elif not network_running and previous_network_running:
        log = {
            "time": current_time,
            "device": rpi1,
            "type": "Network Test",
            "message": "Network test stopped"
        }
        append_to_history_log(log)
    previous_network_running = network_running

    # Read full history from file
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, 'r') as f:
            try:
                full_history = json.load(f)
            except json.JSONDecodeError:
                full_history = []
    else:
        full_history = []

    return jsonify({"histories": full_history[-50:]}), 200

@app.route('/network_metrics', methods=['GET'])
def get_network_metrics():
    if os.path.exists(RESULT_FILE):
        try:
            with open(RESULT_FILE, 'r') as f:
                result = json.load(f)
                summary = result.get("end", {}).get("sum_received", {})
                sender = result.get("end", {}).get("sum_sent", {})

                throughput_mbps = round(summary.get("bits_per_second", 0) / 1_000_000, 2)
                retransmits = sender.get("retransmits", "Unavailable")
                rtt_ms = round(result.get("end", {}).get("streams", [{}])[0].get("receiver", {}).get("mean_rtt", 0) / 1000, 2)
                cpu_utilization = result.get("end", {}).get("cpu_utilization_percent", {}).get("remote_total")
                return jsonify({
                    "client_ip": result.get("start", {}).get("connected", [{}])[0].get("local_host"),
                    "server_ip": result.get("start", {}).get("connected", [{}])[0].get("remote_host"),
                    "timestamp": result.get("start", {}).get("timestamp", {}).get("time"),
                    "sent_bytes": summary.get("bytes", 0),
                    "received_bytes": summary.get("bytes", 0),
                    "cpu_utilization": round(cpu_utilization, 2),
                    "throughput_mbps": throughput_mbps,
                    "retransmits": retransmits,
                    "rtt_ms": rtt_ms
                })
        except Exception as e:
            return jsonify({"status": "error", "message": str(e)})
    else:
        return jsonify({"status": "error", "message": "Result file not found."})

network_running = False
test_result = None
test_lock = threading.Lock()
network_test_process = None

@app.route('/start_network_test', methods=['POST'])
def start_network_test():
    global network_running, test_result, network_test_process

    with test_lock:
        if network_running:
            return jsonify({"status": "error", "message": "Test already running."}), 400

        data = request.json
        target_ip = data.get('target_ip')
        duration = data.get('duration', 20)

        if not target_ip:
            return jsonify({"status": "error", "message": "Missing target IP."}), 400

        def run_test():
            global network_running, test_result, network_test_process
            try:
                command = ["iperf3", "-c", target_ip, "-t", str(duration), "-J"]
                network_test_process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

                stdout, stderr = network_test_process.communicate()

                if network_test_process.returncode == 0:
                    test_result = json.loads(stdout)
                    with open(RESULT_FILE, "w") as result_file:
                        json.dump(test_result, result_file, indent=4)
                else:
                    test_result = {"error": stderr}

            except Exception as e:
                test_result = {"error": str(e)}
            finally:
                with test_lock:
                    network_running = False
                    network_test_process = None

        network_running = True
        test_thread = threading.Thread(target=run_test)
        test_thread.start()

        return jsonify({"status": "started", "message": f"Testing {target_ip} for {duration} seconds."})


@app.route('/stop_network_test', methods=['POST'])
def stop_network_test():
    global network_test_process, network_running
    with test_lock:
        if network_test_process and network_running:
            network_test_process.terminate()
            network_running = False
            return jsonify({"status": "stopped"})
        else:
            return jsonify({"status": "error", "message": "No running test to stop."}), 400


@app.route('/network_test_report', methods=['GET'])
def network_test_report():
    global test_result
    if not test_result:
        return jsonify({"status": "error", "message": "No test result available yet."}), 400
    return jsonify({"status": "ok", "result": test_result})

@app.route("/test_ram")
def test_ram():
    try:
        size_mb = int(request.args.get("size", 100))
        size_bytes = size_mb * 1024 * 1024
        step = 4096  # 4KB step to simulate pages

        # Start time
        start_time = time.time()

        # Allocate memory
        mem_block = bytearray(size_bytes)

        # Write check: write predictable values
        for i in range(0, size_bytes, step):
            mem_block[i] = (i // step) % 256

        # Read check: verify values
        for i in range(0, size_bytes, step):
            if mem_block[i] != (i // step) % 256:
                return jsonify({
                    "status": "error",
                    "message": f"Memory mismatch at offset {i}"
                })

        end_time = time.time()

        # System memory stats
        mem = psutil.virtual_memory()

        return jsonify({
            "status": "success",
            "requested_mb": size_mb,
            "used": mem.total - mem.available,
            "total": mem.total,
            "percent": mem.percent,
            "alloc_speed_sec": round(end_time - start_time, 4),
            "message": "RAM allocation and integrity check passed."
        })

    except MemoryError:
        return jsonify({
            "status": "error",
            "message": "Memory allocation failed. Try a smaller value."
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"An error occurred: {str(e)}"
        })

@app.route('/test_disk', methods=['GET'])
def test_disk():
    try:
        # Get the size of the test file to create
        size_mb = int(request.args.get('size', 100))  # Default to 100MB if not specified
        
        # Create a temporary file to test write speed
        test_file = 'testfile.tmp'
        size_bytes = size_mb * 1024 * 1024

        # Measure write speed
        start_time = time.time()
        with open(test_file, 'wb') as f:
            f.write(os.urandom(size_bytes))  # Write random data
        write_duration = time.time() - start_time

        # Measure read speed
        start_time = time.time()
        with open(test_file, 'rb') as f:
            f.read(size_bytes)  # Read the file data
        read_duration = time.time() - start_time

        # Calculate speed (MB/s)
        write_speed = size_mb / write_duration
        read_speed = size_mb / read_duration

        # Clean up the test file
        os.remove(test_file)

        return jsonify({
            'status': 'success',
            'write_speed': write_speed,
            'read_speed': read_speed,
            'message': 'Disk test completed successfully'
        })

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)